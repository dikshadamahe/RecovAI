"""
HCL Copper Recovery AI – Streamlit Frontend
Run with:  streamlit run app.py
"""

from dotenv import load_dotenv
load_dotenv()
import os
import io
import json
import joblib
import warnings
import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use("Agg")
import seaborn as sns
from datetime import datetime, timedelta
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="HCL Copper Recovery AI",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded",
)

OUTPUT_DIR   = "recovai_output"
DATASET_PATH = "ML_Dataset_Copper_TARGET85.csv"
TARGET       = "Recovery (%)"

FEATURES = [
    "Ore Milled (MT)", "Head Grade (%Cu)", "COPPER IN HEAD (MT)",
    "Feed Rate (MT/h)", "Grinding kWh", "Lime Bags",
    "T Reagent (cc)", "Pine Oil (cc)", "Flotation pH",
    "Milling Running Hours", "SIPX Dose (g/t)", "Frother Dose (g/t)",
    "Depressant Dose (g/t)", "Prev_Recovery (%)",
    "Prev_Feed Rate (MT/h)", "Prev_Head Grade (%Cu)", "Prev_Flotation pH",
    "Roll7_Recovery (%)", "Roll7_Head Grade (%Cu)", "Roll7_Feed Rate (MT/h)",
    "Feed_Condition_Num", "Shift_Num", "Month", "Day_of_Week",
]

REAGENT_COSTS = {          # $/unit (illustrative defaults)
    "SIPX Dose (g/t)":        0.85,
    "Frother Dose (g/t)":     1.10,
    "Depressant Dose (g/t)":  0.65,
}

FEED_CONDITIONS = {0: "Normal", 1: "High Fines", 2: "Low Grade", 3: "Mixed"}
SHIFTS          = {1: "Morning", 2: "Afternoon", 3: "Night"}

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* sidebar gradient */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1a1a2e 0%, #16213e 60%, #0f3460 100%);
}
[data-testid="stSidebar"] * { color: #e0e0e0 !important; }
[data-testid="stSidebar"] .stRadio label { font-size: 0.95rem; }

/* metric cards */
div[data-testid="metric-container"] {
    background: #f8faff;
    border: 1px solid #d6e4ff;
    border-radius: 10px;
    padding: 12px 16px;
}

/* section headers */
.section-header {
    font-size: 1.35rem;
    font-weight: 700;
    color: #1a3a6b;
    border-left: 4px solid #e63946;
    padding-left: 10px;
    margin-bottom: 18px;
}

/* alert boxes */
.alert-critical { background:#fff0f0; border-left:4px solid #e63946; border-radius:6px; padding:12px; margin:6px 0; }
.alert-warning  { background:#fff8e1; border-left:4px solid #ffa000; border-radius:6px; padding:12px; margin:6px 0; }
.alert-normal   { background:#f0fff4; border-left:4px solid #38a169; border-radius:6px; padding:12px; margin:6px 0; }

/* badge */
.badge {
    display:inline-block; padding:3px 10px; border-radius:12px;
    font-size:0.78rem; font-weight:600;
}
.badge-xgb    { background:#dbeafe; color:#1d4ed8; }
.badge-rf     { background:#dcfce7; color:#15803d; }
.badge-linear { background:#fef3c7; color:#92400e; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# LOADERS (cached)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def load_models():
    models = {}
    paths = {
        "XGBoost":       (os.path.join(OUTPUT_DIR, "model_recovery_xgb_clean.json"),    "xgb"),
        "Random Forest": (os.path.join(OUTPUT_DIR, "model_recovery_rf_clean.pkl"),      "pkl"),
        "Linear":        (os.path.join(OUTPUT_DIR, "model_recovery_linear_clean.pkl"),  "pkl"),
    }
    for name, (path, fmt) in paths.items():
        if os.path.exists(path):
            try:
                if fmt == "xgb":
                    from xgboost import XGBRegressor
                    m = XGBRegressor(); m.load_model(path); models[name] = m
                else:
                    models[name] = joblib.load(path)
            except Exception as e:
                st.warning(f"Could not load {name}: {e}")
    return models


@st.cache_data
def load_dataset():
    if not os.path.exists(DATASET_PATH):
        return None
    df = pd.read_csv(DATASET_PATH, header=1)
    df.columns = df.columns.str.strip()
    non_num = {"Date", "Shift", "Estimated Feed Condition", "Source"}
    for c in df.columns:
        if c not in non_num:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
    df = df.sort_values(["Date", "Shift_Num"], na_position="last").reset_index(drop=True)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR NAVIGATION
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ HCL RecovAI")
    st.markdown("**Copper Flotation Intelligence**")
    st.markdown("---")
    page = st.radio(
        "Navigate",
        [
            "🏠 Home",
            "🎯 Recovery Predictor",
            "💊 Reagent Dose Optimizer",
            "🚨 Anomaly & Alert Engine",
            "📊 Shift Performance Dashboard",
            "🔬 Feature Impact Explorer",
            "📝 Shift Report Generator",
            "💬 Ask the Plant",
            "📈 Recovery Trend Forecaster",
        ],
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.markdown("##### Models Loaded")
    models = load_models()
    df_data = load_dataset()
    for mname in ["XGBoost", "Random Forest", "Linear"]:
        icon = "✅" if mname in models else "❌"
        st.markdown(f"{icon} {mname}")
    st.markdown("---")
    st.caption(f"Data: {'✅ Loaded' if df_data is not None else '❌ Not found'}")
    st.caption(f"Rows: {len(df_data):,}" if df_data is not None else "")


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: predict with confidence band
# ─────────────────────────────────────────────────────────────────────────────
def predict_with_ci(model, X_input, model_name, n_bootstraps=50):
    """Returns (point_pred, lower_ci, upper_ci)."""
    point = float(model.predict(X_input)[0])
    if model_name == "Random Forest":
        try:
            preds = np.array([t.predict(X_input)[0] for t in model.estimators_])
            lower = float(np.percentile(preds, 5))
            upper = float(np.percentile(preds, 95))
            return point, lower, upper
        except Exception:
            pass
    # fallback ±2 MAE estimate
    return point, max(0, point - 2.5), min(100, point + 2.5)


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: HOME
# ─────────────────────────────────────────────────────────────────────────────
if page == "🏠 Home":
    st.title("⚙️ HCL Copper Recovery AI Platform")
    st.markdown("**Intelligent decision support for flotation plant operations**")
    st.markdown("---")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        n_models = len(models)
        st.metric("Models Available", n_models, delta="ready" if n_models else "load models")
    with col2:
        n_rows = len(df_data) if df_data is not None else 0
        st.metric("Shift Records", f"{n_rows:,}")
    with col3:
        if df_data is not None and TARGET in df_data.columns:
            avg_rec = df_data[TARGET].mean()
            st.metric("Avg Recovery", f"{avg_rec:.1f}%")
        else:
            st.metric("Avg Recovery", "–")
    with col4:
        st.metric("Features", len(FEATURES))

    st.markdown("---")
    st.markdown("### 📦 Module Overview")

    modules = [
        ("🎯", "Recovery Predictor",        "Predict next-shift recovery % before it starts.",         ["XGBoost / RF", "Confidence Interval", "Shift Form"],       "Core"),
        ("💊", "Reagent Dose Optimizer",    "Find optimal SIPX, frother, depressant for target recovery.",["scipy optimize", "Cost-Aware", "Constrained"],              "Core"),
        ("🚨", "Anomaly & Alert Engine",    "Flag shifts with unusual inputs or likely recovery drop.",  ["Isolation Forest", "Threshold Rules", "Severity Levels"],   "Core"),
        ("📊", "Shift Performance Dashboard","Historical actual vs predicted recovery across all shifts.",["Trend Charts", "Shift Log Table", "Filter"],                "Analysis"),
        ("🔬", "Feature Impact Explorer",   "Which variables actually drive recovery at your plant?",   ["Feature Importance", "Correlation Heatmap", "SHAP info"],    "Analysis"),
        ("📝", "Shift Report Generator",    "Auto-generate a plain-English shift summary.",             ["LLM Summary", "PDF Export", "Manager View"],                 "Analysis"),
        ("💬", "Ask the Plant",             "Natural language Q&A over your shift history.",            ["Groq API", "RAG", "No SQL needed"],                        "Bonus"),
        ("📈", "Recovery Trend Forecaster", "7-day rolling recovery forecast based on planned shifts.", ["Time-Series", "Planned vs Actual", "Weekly Summary"],        "Bonus"),
    ]

    for icon, title, desc, tags, category in modules:
        with st.container():
            c1, c2 = st.columns([0.07, 0.93])
            with c1:
                st.markdown(f"<div style='font-size:2rem;text-align:center'>{icon}</div>", unsafe_allow_html=True)
            with c2:
                tag_html = " ".join(f"<span style='background:#f0f4ff;color:#3366cc;border-radius:10px;padding:2px 8px;font-size:0.75rem;margin-right:4px'>{t}</span>" for t in tags)
                cat_color = "#e63946" if category == "Core" else "#2a9d8f" if category == "Analysis" else "#8338ec"
                st.markdown(
                    f"<b style='font-size:1.05rem'>{title}</b> "
                    f"<span style='background:{cat_color};color:white;border-radius:10px;padding:2px 8px;font-size:0.7rem'>{category}</span><br>"
                    f"<span style='color:#555;font-size:0.88rem'>{desc}</span><br>{tag_html}",
                    unsafe_allow_html=True,
                )
        st.markdown("---")


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: RECOVERY PREDICTOR
# ─────────────────────────────────────────────────────────────────────────────
elif page == "🎯 Recovery Predictor":
    st.markdown('<div class="section-header">🎯 Recovery Predictor</div>', unsafe_allow_html=True)
    st.markdown("Fill in the shift parameters below to predict copper recovery %.")

    if not models:
        st.error("No trained models found. Run the training scripts first and place outputs in `recovai_output/`.")
        st.stop()

    model_choice = st.selectbox("Select Model", list(models.keys()))

    with st.form("predict_form"):
        st.markdown("#### ⛏️ Feed & Milling Inputs")
        c1, c2, c3 = st.columns(3)
        ore_milled   = c1.number_input("Ore Milled (MT)", value=800.0, step=10.0)
        head_grade   = c2.number_input("Head Grade (%Cu)", value=1.5, step=0.05)
        cu_in_head   = c3.number_input("Copper in Head (MT)", value=12.0, step=0.5)
        feed_rate    = c1.number_input("Feed Rate (MT/h)", value=85.0, step=1.0)
        grinding_kwh = c2.number_input("Grinding kWh", value=12000.0, step=100.0)
        milling_hrs  = c3.number_input("Milling Running Hours", value=8.0, step=0.5)

        st.markdown("#### 🧪 Reagent Inputs")
        c1, c2, c3, c4 = st.columns(4)
        ball_mills_running = c1.number_input(
            "Ball Mills Running",
            value=7,
            step=1,
            help="Each ball mill consumes ~11 lime bags on average"
        )
        lime_bags = ball_mills_running * 11
        c1.caption(f"🪨 Lime Bags (auto): {lime_bags}")
        sipx         = c2.number_input("SIPX Dose (g/t)", value=18.0, step=0.5)
        pine_oil     = c3.number_input("Pine Oil (cc)", value=400.0, step=10.0)
        flotation_ph = c4.number_input("Flotation pH", value=10.5, step=0.1)
        frother      = c1.number_input("Frother Dose (g/t)", value=10.0, step=0.5)
        depressant   = c2.number_input("Depressant Dose (g/t)", value=5.0, step=0.5)
        t_reagent    = c3.number_input("T Reagent (cc)", value=1200.0, step=50.0)

        st.markdown("#### 📅 Shift Context")
        c1, c2, c3 = st.columns(3)
        shift_num   = c1.selectbox("Shift", [1, 2, 3], format_func=lambda x: SHIFTS[x])
        feed_cond   = c2.selectbox("Feed Condition", [0, 1, 2, 3], format_func=lambda x: FEED_CONDITIONS[x])
        month       = c3.number_input("Month", value=datetime.today().month, step=1)
        day_of_week = c1.number_input("Day of Week (0=Mon)", value=datetime.today().weekday(), step=1)

        st.markdown("#### 🔁 Previous Shift Values")
        c1, c2, c3, c4 = st.columns(4)
        prev_rec    = c1.number_input("Prev Recovery (%)", value=84.0, step=0.5)
        prev_feed_r = c2.number_input("Prev Feed Rate (MT/h)", value=85.0, step=1.0)
        prev_hg     = c3.number_input("Prev Head Grade (%Cu)", value=1.5, step=0.05)
        prev_ph     = c4.number_input("Prev Flotation pH", value=10.5, step=0.1)

        st.markdown("#### 📊 7-Day Rolling Averages")
        c1, c2, c3 = st.columns(3)
        roll7_rec = c1.number_input("Roll7 Recovery (%)", value=83.5, step=0.5)
        roll7_hg  = c2.number_input("Roll7 Head Grade (%Cu)", value=1.5, step=0.05)
        roll7_fr  = c3.number_input("Roll7 Feed Rate (MT/h)", value=85.0, step=1.0)

        submitted = st.form_submit_button("🔮 Predict Recovery", use_container_width=True)

    if submitted:
        X = pd.DataFrame([[
            ore_milled, head_grade, cu_in_head, feed_rate, grinding_kwh,
            lime_bags, t_reagent, pine_oil, flotation_ph, milling_hrs,
            sipx, frother, depressant,
            prev_rec, prev_feed_r, prev_hg, prev_ph,
            roll7_rec, roll7_hg, roll7_fr,
            float(feed_cond), float(shift_num), float(month), float(day_of_week)
        ]], columns=FEATURES)
        model = models[model_choice]
        point, lower, upper = predict_with_ci(model, X, model_choice)
        st.markdown("---")
        c1, c2, c3 = st.columns(3)
        c1.metric("Predicted Recovery", f"{point:.2f}%")
        c2.metric("90% CI Lower",       f"{lower:.2f}%")
        c3.metric("90% CI Upper",       f"{upper:.2f}%")

        # ── Metallurgical KPIs ──────────────────────────────────────────────
        st.markdown("#### ⚗️ Metallurgical Outputs")
        cu_in_head_calc = (head_grade / 100) * ore_milled
        CONC_GRADE_ASSUMED = 28.0
        cu_recovered_mt   = cu_in_head_calc * (point / 100)
        final_concentrate = cu_recovered_mt / (CONC_GRADE_ASSUMED / 100)
        recovery_kpi      = point
        conc_grade_kpi    = CONC_GRADE_ASSUMED

        kpi1, kpi2, kpi3 = st.columns(3)
        kpi1.metric(
            "🏭 Final Concentrate Generated",
            f"{final_concentrate:.1f} MT",
            help="Estimated dry concentrate mass this shift"
        )
        kpi2.metric(
            "📈 Recovery (%)",
            f"{recovery_kpi:.2f}%",
            delta=f"{recovery_kpi - 85:.2f}% vs target",
            delta_color="normal"
        )
        kpi3.metric(
            "🔬 Concentrate Grade (%Cu)",
            f"{conc_grade_kpi:.1f}%",
            help="Target concentrate grade benchmark (28%Cu)"
        )
        st.caption(
            f"💡 Based on: Cu in Head = {cu_in_head_calc:.2f} MT · "
            f"Cu Recovered = {cu_recovered_mt:.2f} MT · "
            f"Concentrate Grade assumption = {CONC_GRADE_ASSUMED}%Cu"
        )
        st.markdown("---")

        status_color = "#38a169" if point >= 85 else "#ffa000" if point >= 80 else "#e63946"
        status_label = "✅ Above Target (85%)" if point >= 85 else "⚠️ Near Target" if point >= 80 else "❌ Below Target"
        st.markdown(
            f"<div style='background:{status_color}22;border-left:4px solid {status_color};border-radius:6px;padding:14px;font-size:1.1rem'>"
            f"<b>{status_label}</b> — Model: <b>{model_choice}</b> | CI: [{lower:.1f}%, {upper:.1f}%]"
            f"</div>",
            unsafe_allow_html=True,
        )

        # Gauge chart
        fig, ax = plt.subplots(figsize=(5, 2.5))
        ax.barh([0], [100], color="#eeeeee", height=0.4)
        color = "#e63946" if point < 80 else "#ffa000" if point < 85 else "#38a169"
        ax.barh([0], [point], color=color, height=0.4)
        ax.axvline(x=85, color="navy", linestyle="--", linewidth=1.5, label="Target 85%")
        ax.set_xlim(70, 100)
        ax.set_yticks([])
        ax.set_xlabel("Recovery (%)")
        ax.set_title("Recovery Gauge")
        ax.legend(fontsize=8)
        ax.fill_betweenx([-0.2, 0.2], lower, upper, color=color, alpha=0.2, label="90% CI")
        plt.tight_layout()
        st.pyplot(fig)
        plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: REAGENT DOSE OPTIMIZER
# ─────────────────────────────────────────────────────────────────────────────
elif page == "💊 Reagent Dose Optimizer":
    st.markdown('<div class="section-header">💊 Reagent Dose Optimizer</div>', unsafe_allow_html=True)
    st.markdown("Find the minimum-cost reagent doses that achieve a target recovery, given fixed shift conditions.")

    if not models:
        st.error("No models loaded. Train models first.")
        st.stop()

    model_choice = st.selectbox("Model for optimization", list(models.keys()))
    target_rec   = st.slider("Target Recovery (%)", 75.0, 98.0, 85.0, step=0.5)

    st.markdown("#### Fixed Shift Conditions")
    c1, c2, c3 = st.columns(3)
    ore_milled   = c1.number_input("Ore Milled (MT)", value=800.0)
    head_grade   = c2.number_input("Head Grade (%Cu)", value=1.5, step=0.05)
    cu_in_head   = c3.number_input("Copper in Head (MT)", value=12.0)
    feed_rate    = c1.number_input("Feed Rate (MT/h)", value=85.0)
    grinding_kwh = c2.number_input("Grinding kWh", value=12000.0)
    flotation_ph = c3.number_input("Flotation pH", value=10.5, step=0.1)
    ball_mills_opt = c1.number_input("Ball Mills Running", value=7, step=1, help="Each mill = 11 lime bags")
    lime_bags      = ball_mills_opt * 11
    c1.caption(f"🪨 Lime Bags (auto): {lime_bags}")
    milling_hrs  = c2.number_input("Milling Hrs", value=8.0)
    t_reagent    = c3.number_input("T Reagent (cc)", value=1200.0)
    pine_oil     = c1.number_input("Pine Oil (cc)", value=400.0)
    prev_rec     = c2.number_input("Prev Recovery (%)", value=84.0)
    prev_feed_r  = c3.number_input("Prev Feed Rate", value=85.0)
    prev_hg      = c1.number_input("Prev Head Grade", value=1.5, step=0.05)
    prev_ph      = c2.number_input("Prev Flotation pH", value=10.5, step=0.1)
    roll7_rec    = c3.number_input("Roll7 Recovery", value=83.5)
    roll7_hg     = c1.number_input("Roll7 Head Grade", value=1.5, step=0.05)
    roll7_fr     = c2.number_input("Roll7 Feed Rate", value=85.0)
    shift_num    = c3.selectbox("Shift", [1, 2, 3], format_func=lambda x: SHIFTS[x])
    feed_cond    = c1.selectbox("Feed Condition", [0, 1, 2, 3], format_func=lambda x: FEED_CONDITIONS[x])
    month        = c2.number_input("Month", 1, 12, datetime.today().month)
    dow          = c3.number_input("Day of Week", 0, 6, datetime.today().weekday())

    st.markdown("#### Reagent Bounds & Costs")
    rc1, rc2, rc3 = st.columns(3)
    sipx_min, sipx_max       = rc1.slider("SIPX (g/t) range",       0.0,  50.0, (5.0, 35.0))
    frother_min, frother_max = rc2.slider("Frother (g/t) range",    0.0,  30.0, (3.0, 20.0))
    dep_min, dep_max         = rc3.slider("Depressant (g/t) range", 0.0,  20.0, (1.0, 12.0))
    sipx_cost    = rc1.number_input("SIPX cost ($/g/t)", value=0.85, step=0.05)
    frother_cost = rc2.number_input("Frother cost ($/g/t)", value=1.10, step=0.05)
    dep_cost     = rc3.number_input("Depressant cost ($/g/t)", value=0.65, step=0.05)

    if st.button("🔍 Optimize Doses", use_container_width=True):
        model = models[model_choice]
        fixed_vals = [
            ore_milled, head_grade, cu_in_head, feed_rate, grinding_kwh,
            lime_bags, t_reagent, pine_oil, flotation_ph, milling_hrs,
            prev_rec, prev_feed_r, prev_hg, prev_ph,
            roll7_rec, roll7_hg, roll7_fr,
            float(feed_cond), float(shift_num), float(month), float(dow)
        ]
        # FEATURES order: ...SIPX[10], Frother[11], Depressant[12]...
        def make_row(sipx, frother, dep):
            vals = [
                ore_milled, head_grade, cu_in_head, feed_rate, grinding_kwh,
                lime_bags, t_reagent, pine_oil, flotation_ph, milling_hrs,
                sipx, frother, dep,
                prev_rec, prev_feed_r, prev_hg, prev_ph,
                roll7_rec, roll7_hg, roll7_fr,
                float(feed_cond), float(shift_num), float(month), float(dow)
            ]
            return pd.DataFrame([vals], columns=FEATURES)

        def cost_fn(x):
            sipx, frother, dep = x
            cost = sipx * sipx_cost + frother * frother_cost + dep * dep_cost
            pred = model.predict(make_row(sipx, frother, dep))[0]
            penalty = 10000 * max(0, target_rec - pred) ** 2
            return cost + penalty

        bounds = [(sipx_min, sipx_max), (frother_min, frother_max), (dep_min, dep_max)]
        x0 = [(sipx_min + sipx_max) / 2, (frother_min + frother_max) / 2, (dep_min + dep_max) / 2]
        result = minimize(cost_fn, x0, bounds=bounds, method="L-BFGS-B")
        opt_sipx, opt_frother, opt_dep = result.x
        opt_pred = float(model.predict(make_row(opt_sipx, opt_frother, opt_dep))[0])
        opt_cost = opt_sipx * sipx_cost + opt_frother * frother_cost + opt_dep * dep_cost

        st.success("Optimization complete!")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Optimal SIPX",       f"{opt_sipx:.2f} g/t")
        c2.metric("Optimal Frother",    f"{opt_frother:.2f} g/t")
        c3.metric("Optimal Depressant", f"{opt_dep:.2f} g/t")
        c4.metric("Predicted Recovery", f"{opt_pred:.2f}%")

        total_cost = opt_cost
        status = "✅ Target Met" if opt_pred >= target_rec else "⚠️ Target Missed"
        st.info(f"{status} | Total reagent cost index: **{total_cost:.2f}** $/t ore")

        fig, ax = plt.subplots(figsize=(6, 3))
        reagents = ["SIPX", "Frother", "Depressant"]
        values   = [opt_sipx, opt_frother, opt_dep]
        colors   = ["#3b82f6", "#10b981", "#f59e0b"]
        ax.barh(reagents, values, color=colors)
        ax.set_xlabel("Dose (g/t)")
        ax.set_title("Optimized Reagent Doses")
        for i, v in enumerate(values):
            ax.text(v + 0.1, i, f"{v:.2f}", va="center", fontsize=9)
        plt.tight_layout()
        st.pyplot(fig)
        plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: ANOMALY & ALERT ENGINE
# ─────────────────────────────────────────────────────────────────────────────
elif page == "🚨 Anomaly & Alert Engine":
    st.markdown('<div class="section-header">🚨 Anomaly & Alert Engine</div>', unsafe_allow_html=True)
    st.markdown("Flag shifts where inputs look unusual or recovery is likely to drop.")

    if df_data is None:
        st.error(f"Dataset not found at `{DATASET_PATH}`.")
        st.stop()

    df = df_data.copy()
    avail_features = [f for f in FEATURES if f in df.columns]
    X_all = df[avail_features].dropna()

    try:
        from sklearn.ensemble import IsolationForest
        contamination = st.slider("Expected anomaly fraction", 0.01, 0.20, 0.05, step=0.01)
        iso = IsolationForest(contamination=contamination, random_state=42)
        scores = iso.fit_predict(X_all)
        anomaly_df = df.loc[X_all.index].copy()
        anomaly_df["Anomaly"] = scores
        anomaly_df["AnomalyScore"] = iso.decision_function(X_all)
        anomaly_df["IsAnomaly"] = anomaly_df["Anomaly"] == -1
    except Exception as e:
        st.error(f"Isolation Forest failed: {e}")
        st.stop()

    # Threshold alerts
    def classify_severity(row):
        if row.get("IsAnomaly", False):
            score = row.get("AnomalyScore", 0)
            if score < -0.15:
                return "Critical"
            return "Warning"
        if TARGET in row and pd.notna(row[TARGET]) and row[TARGET] < 80:
            return "Warning"
        return "Normal"

    anomaly_df["Severity"] = anomaly_df.apply(classify_severity, axis=1)

    n_critical = (anomaly_df["Severity"] == "Critical").sum()
    n_warning  = (anomaly_df["Severity"] == "Warning").sum()
    n_normal   = (anomaly_df["Severity"] == "Normal").sum()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Shifts", len(anomaly_df))
    c2.metric("🔴 Critical",  n_critical)
    c3.metric("🟡 Warning",   n_warning)
    c4.metric("🟢 Normal",    n_normal)

    severity_filter = st.multiselect("Filter by severity", ["Critical", "Warning", "Normal"], default=["Critical", "Warning"])
    filtered = anomaly_df[anomaly_df["Severity"].isin(severity_filter)]

    display_cols = ["Date", "Shift", TARGET, "Head Grade (%Cu)", "Feed Rate (MT/h)", "Severity", "AnomalyScore"]
    display_cols = [c for c in display_cols if c in filtered.columns]

    def color_severity(val):
        colors_map = {"Critical": "background-color:#ffe4e4", "Warning": "background-color:#fff8e1", "Normal": "background-color:#f0fff4"}
        return colors_map.get(val, "")

    st.dataframe(
        filtered[display_cols].tail(200).style.applymap(color_severity, subset=["Severity"]),
        use_container_width=True, height=350
    )

    # Anomaly score over time
    if "Date" in anomaly_df.columns:
        fig, ax = plt.subplots(figsize=(12, 3))
        ax.plot(anomaly_df["Date"], anomaly_df["AnomalyScore"], color="#aaa", linewidth=0.8, label="Score")
        anomalies = anomaly_df[anomaly_df["IsAnomaly"]]
        ax.scatter(anomalies["Date"], anomalies["AnomalyScore"], color="#e63946", s=20, zorder=5, label="Anomaly")
        ax.axhline(0, linestyle="--", color="#999", linewidth=0.8)
        ax.set_xlabel("Date")
        ax.set_ylabel("Anomaly Score")
        ax.set_title("Isolation Forest Anomaly Scores Over Time")
        ax.legend()
        plt.tight_layout()
        st.pyplot(fig)
        plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: SHIFT PERFORMANCE DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
elif page == "📊 Shift Performance Dashboard":
    st.markdown('<div class="section-header">📊 Shift Performance Dashboard</div>', unsafe_allow_html=True)

    if df_data is None:
        st.error(f"Dataset not found at `{DATASET_PATH}`.")
        st.stop()

    df = df_data.copy()

    # Filters
    fc1, fc2, fc3 = st.columns(3)
    months = sorted(df["Date"].dt.month.dropna().unique().astype(int).tolist())
    sel_months = fc1.multiselect("Month", months, default=months)
    if "Estimated Feed Condition" in df.columns:
        conds = df["Estimated Feed Condition"].dropna().unique().tolist()
        sel_conds = fc2.multiselect("Feed Condition", conds, default=conds)
        df = df[df["Estimated Feed Condition"].isin(sel_conds)]
    shifts_list = sorted(df["Shift_Num"].dropna().unique().astype(int).tolist()) if "Shift_Num" in df.columns else [1, 2, 3]
    sel_shifts = fc3.multiselect("Shift", shifts_list, default=shifts_list)

    df = df[df["Date"].dt.month.isin(sel_months)]
    if "Shift_Num" in df.columns:
        df = df[df["Shift_Num"].isin(sel_shifts)]

    df_target = df[df[TARGET].notna()].copy()

    if models:
        model_choice = st.selectbox("Model for prediction overlay", list(models.keys()))
        model = models[model_choice]
        avail = [f for f in FEATURES if f in df_target.columns]
        Xp = df_target[avail].copy()
        for c in FEATURES:
            if c not in Xp.columns:
                Xp[c] = 0.0
        Xp = Xp[FEATURES].fillna(0)
        df_target["Predicted"] = model.predict(Xp)
    else:
        df_target["Predicted"] = np.nan

    # Summary metrics
    cm1, cm2, cm3, cm4 = st.columns(4)
    cm1.metric("Avg Actual Recovery",    f"{df_target[TARGET].mean():.2f}%")
    cm2.metric("Shifts Above 85%",       f"{(df_target[TARGET] >= 85).sum()} / {len(df_target)}")
    if df_target["Predicted"].notna().any():
        cm3.metric("Avg Predicted",      f"{df_target['Predicted'].mean():.2f}%")
        mae = (df_target[TARGET] - df_target["Predicted"]).abs().mean()
        cm4.metric("Mean Abs Error",     f"{mae:.2f}%")

    # Trend chart
    fig, ax = plt.subplots(figsize=(14, 4))
    ax.plot(df_target["Date"], df_target[TARGET], color="#3b82f6", linewidth=1, label="Actual", alpha=0.8)
    if df_target["Predicted"].notna().any():
        ax.plot(df_target["Date"], df_target["Predicted"], color="#f59e0b", linewidth=1, linestyle="--", label="Predicted", alpha=0.8)
    ax.axhline(85, color="#e63946", linestyle=":", linewidth=1.5, label="Target 85%")
    ax.set_xlabel("Date")
    ax.set_ylabel("Recovery (%)")
    ax.set_title("Actual vs Predicted Recovery Over Time")
    ax.legend()
    ax.grid(alpha=0.2)
    plt.tight_layout()
    st.pyplot(fig)
    plt.close()

    # Monthly box plot
    if "Date" in df_target.columns:
        df_target["Month_Name"] = df_target["Date"].dt.strftime("%b-%y")
        months_order = df_target.sort_values("Date")["Month_Name"].unique().tolist()
        fig2, ax2 = plt.subplots(figsize=(14, 4))
        month_data = [df_target[df_target["Month_Name"] == m][TARGET].dropna().values for m in months_order]
        ax2.boxplot(month_data, labels=months_order, patch_artist=True,
                    boxprops=dict(facecolor="#dbeafe"), medianprops=dict(color="#1d4ed8", linewidth=2))
        ax2.axhline(85, color="#e63946", linestyle=":", linewidth=1.5, label="Target 85%")
        ax2.set_title("Monthly Recovery Distribution")
        ax2.set_ylabel("Recovery (%)")
        ax2.tick_params(axis='x', rotation=45)
        ax2.legend()
        plt.tight_layout()
        st.pyplot(fig2)
        plt.close()

    # Shift log table
    st.markdown("#### 📋 Shift Log (last 50 shifts)")
    display_cols = ["Date", "Shift", TARGET, "Head Grade (%Cu)", "Feed Rate (MT/h)", "SIPX Dose (g/t)", "Frother Dose (g/t)"]
    if "Predicted" in df_target.columns:
        display_cols.append("Predicted")
    display_cols = [c for c in display_cols if c in df_target.columns]
    st.dataframe(df_target[display_cols].tail(50), use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: FEATURE IMPACT EXPLORER
# ─────────────────────────────────────────────────────────────────────────────
elif page == "🔬 Feature Impact Explorer":
    st.markdown('<div class="section-header">🔬 Feature Impact Explorer</div>', unsafe_allow_html=True)
    st.markdown("Understand which variables drive copper recovery at your plant.")

    tab1, tab2, tab3 = st.tabs(["📊 Feature Importance", "🌡️ Correlation Heatmap", "📈 Partial Dependence"])

    with tab1:
        # Show pre-generated plots if available
        img_paths = {
            "XGBoost Feature Importance":       os.path.join(OUTPUT_DIR, "xgb_clean_feature_importance.png"),
            "RF Feature Importance":            os.path.join(OUTPUT_DIR, "rf_clean_feature_importance.png"),
            "RF Permutation Importance":        os.path.join(OUTPUT_DIR, "rf_clean_permutation_importance.png"),
            "SHAP Summary (Recovery)":          os.path.join(OUTPUT_DIR, "plot_shap_recovery.png"),
        }
        for title, path in img_paths.items():
            if os.path.exists(path):
                st.markdown(f"**{title}**")
                st.image(path, use_container_width=True)

        # Live importance from loaded models
        if models and df_data is not None:
            st.markdown("#### Live Feature Importance")
            model_choice = st.selectbox("Choose model", list(models.keys()), key="fi_model")
            model = models[model_choice]
            if hasattr(model, "feature_importances_"):
                imp = pd.Series(model.feature_importances_, index=FEATURES).sort_values(ascending=False)
                fig, ax = plt.subplots(figsize=(8, 6))
                imp.head(15).sort_values().plot(kind="barh", ax=ax, color="#3b82f6")
                ax.set_title(f"Top 15 Features – {model_choice}")
                ax.set_xlabel("Importance")
                plt.tight_layout()
                st.pyplot(fig)
                plt.close()

    with tab2:
        if df_data is None:
            st.warning("Dataset not loaded.")
        else:
            numeric_cols = [f for f in FEATURES if f in df_data.columns] + [TARGET]
            subset = df_data[numeric_cols].dropna()
            top_n = st.slider("Top N features", 5, 20, 12)
            corr_with_target = subset.corr()[TARGET].abs().sort_values(ascending=False)
            top_features = corr_with_target.index[:top_n + 1].tolist()
            corr_matrix = subset[top_features].corr()
            fig, ax = plt.subplots(figsize=(10, 8))
            sns.heatmap(corr_matrix, annot=True, fmt=".2f", cmap="coolwarm",
                        center=0, ax=ax, linewidths=0.5, annot_kws={"size": 7})
            ax.set_title(f"Correlation Heatmap — Top {top_n} Features vs {TARGET}")
            plt.tight_layout()
            st.pyplot(fig)
            plt.close()

    with tab3:
        if df_data is None or not models:
            st.warning("Need dataset and models for partial dependence plots.")
        else:
            model_choice = st.selectbox("Model", list(models.keys()), key="pdp_model")
            feature = st.selectbox("Feature to plot", [f for f in FEATURES if f in df_data.columns])
            model = models[model_choice]
            avail = [f for f in FEATURES if f in df_data.columns]
            Xp = df_data[avail].dropna().copy()
            for c in FEATURES:
                if c not in Xp.columns:
                    Xp[c] = 0.0
            Xp = Xp[FEATURES]
            if len(Xp) > 500:
                Xp = Xp.sample(500, random_state=42)
            grid = np.linspace(Xp[feature].quantile(0.05), Xp[feature].quantile(0.95), 40)
            pdp_vals = []
            for val in grid:
                Xtemp = Xp.copy()
                Xtemp[feature] = val
                pdp_vals.append(model.predict(Xtemp).mean())
            fig, ax = plt.subplots(figsize=(8, 4))
            ax.plot(grid, pdp_vals, color="#3b82f6", linewidth=2)
            ax.set_xlabel(feature)
            ax.set_ylabel("Avg Predicted Recovery (%)")
            ax.set_title(f"Partial Dependence: {feature}")
            ax.grid(alpha=0.2)
            plt.tight_layout()
            st.pyplot(fig)
            plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: SHIFT REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────
elif page == "📝 Shift Report Generator":
    st.markdown('<div class="section-header">📝 Shift Report Generator</div>', unsafe_allow_html=True)
    st.markdown("Auto-generate a plain-English shift summary. Uses Groq AI (Llama 3).")

    if df_data is None:
        st.error("Dataset not loaded.")
        st.stop()

    df = df_data.copy()
    df["ShiftLabel"] = df["Date"].dt.strftime("%Y-%m-%d") + " " + df["Shift"].fillna("").astype(str)
    shift_labels = df["ShiftLabel"].dropna().tolist()

    selected_shift_label = st.selectbox("Select Shift", shift_labels[-100:])
    view = st.radio("Report View", ["Operator", "Manager"], horizontal=True)
    use_ai = st.checkbox("✨ Enhance with AI (requires Groq API)", value=False)

    row = df[df["ShiftLabel"] == selected_shift_label].iloc[0]

    if st.button("📄 Generate Report", use_container_width=True):
        recovery = row.get(TARGET, "N/A")
        head_grade = row.get("Head Grade (%Cu)", "N/A")
        feed_rate  = row.get("Feed Rate (MT/h)", "N/A")
        sipx       = row.get("SIPX Dose (g/t)", "N/A")
        frother    = row.get("Frother Dose (g/t)", "N/A")
        ph         = row.get("Flotation pH", "N/A")

        status = "above target ✅" if pd.notna(recovery) and recovery >= 85 else "below target ⚠️"

        report_text = f"""
**SHIFT REPORT — {selected_shift_label}**
View: {view}

---
**Recovery:** {recovery:.2f}% ({status}) | **Target:** 85%
**Head Grade:** {head_grade} %Cu | **Feed Rate:** {feed_rate} MT/h
**SIPX Dose:** {sipx} g/t | **Frother Dose:** {frother} g/t | **Flotation pH:** {ph}
"""
        if view == "Operator":
            report_text += "\n**Recommendations:** Monitor reagent dosing. Ensure pH stays in 10.2–11.2 range."
        else:
            report_text += "\n**Summary:** Recovery performance relative to target. Review input conditions for the next shift."

        if use_ai:
            st.info("Calling Groq AI for AI-enhanced report...")
            try:
                from groq import Groq
                groq_client = Groq(api_key=os.environ.get("GROQ_API_KEY", ""))
                prompt = (
                    f"You are a copper flotation expert. Write a concise {view.lower()}-level shift report "
                    f"based on these metrics:\n"
                    f"- Recovery: {recovery:.2f}% (target 85%)\n"
                    f"- Head Grade: {head_grade} %Cu\n"
                    f"- Feed Rate: {feed_rate} MT/h\n"
                    f"- SIPX: {sipx} g/t, Frother: {frother} g/t, pH: {ph}\n"
                    f"Keep it under 150 words. {'Use simple operational language.' if view=='Operator' else 'Include a performance summary and key action items.'}"
                )
                chat_completion = groq_client.chat.completions.create(
                    messages=[{"role": "user", "content": prompt}],
                    model="llama-3.3-70b-versatile",
                    max_tokens=300,
                )
                ai_text = chat_completion.choices[0].message.content
                if ai_text:
                    report_text += f"\n\n---\n**AI Analysis:**\n{ai_text}"
            except Exception as e:
                st.warning(f"Groq API call failed: {e}. Showing basic report.")

        st.markdown(report_text)

        # ── XLSX Export ─────────────────────────────────────────────────────
        def build_report_xlsx(row, recovery, head_grade, feed_rate, sipx, frother, ph,
                               selected_shift_label, view, status, ai_text=None):
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, numbers)
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Shift Report"

            # ── colour palette ───────────────────────────────────────────────
            DARK_BLUE   = "1A3A6B"
            MID_BLUE    = "3B82F6"
            LIGHT_BLUE  = "DBEAFE"
            GREEN_BG    = "F0FFF4"
            GREEN_BD    = "38A169"
            AMBER_BG    = "FFF8E1"
            AMBER_BD    = "FFA000"
            RED_BG      = "FFF0F0"
            RED_BD      = "E63946"
            WHITE       = "FFFFFF"
            HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=13)
            TITLE_FONT  = Font(name="Arial", bold=True, color=DARK_BLUE, size=11)
            LABEL_FONT  = Font(name="Arial", bold=True, color="444444", size=10)
            VALUE_FONT  = Font(name="Arial", color="111111", size=10)
            SMALL_FONT  = Font(name="Arial", color="666666", size=9, italic=True)
            thin = Side(style="thin", color="CCCCCC")
            std_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

            def hfill(color): return PatternFill("solid", fgColor=color)

            # col widths
            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 28
            ws.column_dimensions["D"].width = 22

            # ── Header banner ────────────────────────────────────────────────
            ws.merge_cells("A1:D1")
            ws["A1"] = "⚙  HCL COPPER RECOVERY AI — SHIFT REPORT"
            ws["A1"].font   = HEADER_FONT
            ws["A1"].fill   = hfill(DARK_BLUE)
            ws["A1"].alignment = center
            ws.row_dimensions[1].height = 30

            ws.merge_cells("A2:D2")
            ws["A2"] = f"Shift: {selected_shift_label}   |   View: {view}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws["A2"].font      = Font(name="Arial", color="AAAAAA", size=9, italic=True)
            ws["A2"].fill      = hfill("1E2D4F")
            ws["A2"].alignment = center
            ws.row_dimensions[2].height = 18

            # ── Section helper ───────────────────────────────────────────────
            def section_header(row_num, title):
                ws.merge_cells(f"A{row_num}:D{row_num}")
                ws[f"A{row_num}"] = title
                ws[f"A{row_num}"].font      = Font(name="Arial", bold=True, color=DARK_BLUE, size=11)
                ws[f"A{row_num}"].fill      = hfill(LIGHT_BLUE)
                ws[f"A{row_num}"].alignment = left
                ws[f"A{row_num}"].border    = Border(bottom=Side(style="medium", color=MID_BLUE))
                ws.row_dimensions[row_num].height = 22

            def kv(row_num, label, value, unit="", note="", row_fill=None):
                ws[f"A{row_num}"] = label
                ws[f"A{row_num}"].font      = LABEL_FONT
                ws[f"A{row_num}"].alignment = left
                ws[f"A{row_num}"].border    = std_border
                ws[f"B{row_num}"] = value
                ws[f"B{row_num}"].font      = VALUE_FONT
                ws[f"B{row_num}"].alignment = center
                ws[f"B{row_num}"].border    = std_border
                ws[f"C{row_num}"] = unit
                ws[f"C{row_num}"].font      = SMALL_FONT
                ws[f"C{row_num}"].alignment = center
                ws[f"C{row_num}"].border    = std_border
                ws[f"D{row_num}"] = note
                ws[f"D{row_num}"].font      = SMALL_FONT
                ws[f"D{row_num}"].alignment = left
                ws[f"D{row_num}"].border    = std_border
                if row_fill:
                    for col in ["A","B","C","D"]:
                        ws[f"{col}{row_num}"].fill = hfill(row_fill)
                ws.row_dimensions[row_num].height = 18

            # ── Recovery Status Banner ────────────────────────────────────────
            r = 4
            ws.merge_cells(f"A{r}:D{r}")
            if pd.notna(recovery) and recovery >= 85:
                status_txt = f"✅  Recovery: {recovery:.2f}%  —  ABOVE TARGET (85%)"
                bg, bd = GREEN_BG, GREEN_BD
            elif pd.notna(recovery) and recovery >= 80:
                status_txt = f"⚠️  Recovery: {recovery:.2f}%  —  NEAR TARGET (85%)"
                bg, bd = AMBER_BG, AMBER_BD
            else:
                rv = f"{recovery:.2f}%" if pd.notna(recovery) else "N/A"
                status_txt = f"❌  Recovery: {rv}  —  BELOW TARGET (85%)"
                bg, bd = RED_BG, RED_BD
            ws[f"A{r}"] = status_txt
            ws[f"A{r}"].font      = Font(name="Arial", bold=True, color=bd, size=12)
            ws[f"A{r}"].fill      = hfill(bg)
            ws[f"A{r}"].alignment = center
            ws[f"A{r}"].border    = Border(left=Side(style="thick", color=bd),
                                            right=Side(style="thin", color="CCCCCC"),
                                            top=Side(style="thin", color="CCCCCC"),
                                            bottom=Side(style="thin", color="CCCCCC"))
            ws.row_dimensions[r].height = 26

            # ── Column headers ────────────────────────────────────────────────
            r = 6
            ws.row_dimensions[r].height = 20
            for col, hdr in zip(["A","B","C","D"],
                                 ["Parameter", "Value", "Unit / Range", "Notes"]):
                ws[f"{col}{r}"] = hdr
                ws[f"{col}{r}"].font      = Font(name="Arial", bold=True, color=WHITE, size=10)
                ws[f"{col}{r}"].fill      = hfill(MID_BLUE)
                ws[f"{col}{r}"].alignment = center
                ws[f"{col}{r}"].border    = std_border

            # ── Section 1: Recovery & Grade ───────────────────────────────────
            r = 7;  section_header(r, "1  |  Recovery & Grade")
            r = 8;  kv(r, "Recovery (%)",      f"{recovery:.2f}" if pd.notna(recovery) else "N/A", "%", "Target: 85%", "F9FAFB")
            r = 9;  kv(r, "Head Grade",        f"{head_grade}"  if pd.notna(head_grade) else "N/A", "%Cu")
            r = 10; kv(r, "Target Recovery",   "85.00", "%", "Plant benchmark", "F9FAFB")
            delta = round(recovery - 85, 2) if pd.notna(recovery) else "N/A"
            r = 11; kv(r, "Delta vs Target",   delta, "%", "Positive = above target")

            # ── Section 2: Process Inputs ─────────────────────────────────────
            r = 13; section_header(r, "2  |  Process Inputs")
            r = 14; kv(r, "Feed Rate",         f"{feed_rate}"  if pd.notna(feed_rate)  else "N/A", "MT/h", "",           "F9FAFB")
            r = 15; kv(r, "Ore Milled",        f"{row.get('Ore Milled (MT)', 'N/A')}", "MT")
            r = 16; kv(r, "Flotation pH",      f"{ph}"         if pd.notna(ph)         else "N/A", "",  "Optimal: 10.2–11.2", "F9FAFB")
            r = 17; kv(r, "Milling kWh",       f"{row.get('Grinding kWh', 'N/A')}", "kWh")
            r = 18; kv(r, "Milling Hours",     f"{row.get('Milling Running Hours', 'N/A')}", "hrs", "", "F9FAFB")

            # ── Section 3: Reagent Dosing ─────────────────────────────────────
            r = 20; section_header(r, "3  |  Reagent Dosing")
            r = 21; kv(r, "SIPX Dose",         f"{sipx}"       if pd.notna(sipx)       else "N/A", "g/t", "",           "F9FAFB")
            r = 22; kv(r, "Frother Dose",      f"{frother}"    if pd.notna(frother)    else "N/A", "g/t")
            r = 23; kv(r, "Depressant Dose",   f"{row.get('Depressant Dose (g/t)', 'N/A')}", "g/t", "",                 "F9FAFB")
            r = 24; kv(r, "T Reagent",         f"{row.get('T Reagent (cc)', 'N/A')}", "cc")
            r = 25; kv(r, "Pine Oil",          f"{row.get('Pine Oil (cc)', 'N/A')}", "cc", "",                          "F9FAFB")
            r = 26; kv(r, "Lime Bags",         f"{row.get('Lime Bags', 'N/A')}", "bags")

            # ── Section 4: Shift Context ──────────────────────────────────────
            r = 28; section_header(r, "4  |  Shift Context")
            r = 29; kv(r, "Shift",             f"{row.get('Shift', 'N/A')}", "", "",                                    "F9FAFB")
            r = 30; kv(r, "Feed Condition",    f"{row.get('Estimated Feed Condition', 'N/A')}")
            r = 31; kv(r, "Copper in Head",    f"{row.get('COPPER IN HEAD (MT)', 'N/A')}", "MT", "",                    "F9FAFB")
            r = 32; kv(r, "Prev Recovery",     f"{row.get('Prev_Recovery (%)', 'N/A')}", "%")
            r = 33; kv(r, "7-Day Avg Recovery",f"{row.get('Roll7_Recovery (%)', 'N/A')}", "%", "Rolling 7-shift avg",   "F9FAFB")

            # ── Section 5: Recommendations ────────────────────────────────────
            r = 35; section_header(r, "5  |  Recommendations")
            if view == "Operator":
                recs = [
                    ("Monitor reagent dosing", "Keep SIPX and frother within optimal range"),
                    ("Check flotation pH",     "Ensure pH stays between 10.2 and 11.2"),
                    ("Feed rate control",      "Maintain stable feed rate to avoid recovery dips"),
                ]
            else:
                recs = [
                    ("Performance review",     f"Recovery {recovery:.2f}% vs 85% target — delta {delta}%"),
                    ("Reagent cost check",     "Review SIPX/frother consumption vs budget"),
                    ("Next shift planning",    "Verify feed condition and grade forecast for next shift"),
                ]
            for i, (action, detail) in enumerate(recs):
                r = 36 + i
                ws[f"A{r}"] = action
                ws[f"A{r}"].font = LABEL_FONT
                ws[f"A{r}"].border = std_border
                ws[f"A{r}"].alignment = left
                ws.merge_cells(f"B{r}:D{r}")
                ws[f"B{r}"] = detail
                ws[f"B{r}"].font = VALUE_FONT
                ws[f"B{r}"].border = std_border
                ws[f"B{r}"].alignment = left
                if i % 2 == 0:
                    for col in ["A","B"]:
                        ws[f"{col}{r}"].fill = hfill("F9FAFB")
                ws.row_dimensions[r].height = 18

            # ── Section 6: AI Analysis (if present) ──────────────────────────
            if ai_text:
                r = 40; section_header(r, "6  |  AI Analysis (Groq / Llama 3)")
                r = 41
                ws.merge_cells(f"A{r}:D{r+8}")
                ws[f"A{r}"] = ai_text
                ws[f"A{r}"].font = Font(name="Arial", color="1A3A6B", size=10)
                ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws[f"A{r}"].fill = hfill("EEF4FF")
                ws[f"A{r}"].border = Border(
                    left=Side(style="thick", color=MID_BLUE),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"),
                )
                for rr in range(r, r+9):
                    ws.row_dimensions[rr].height = 16

            # ── Footer ────────────────────────────────────────────────────────
            last = 51 if ai_text else 42
            ws.merge_cells(f"A{last}:D{last}")
            ws[f"A{last}"] = f"Generated by HCL RecovAI  •  {datetime.now().strftime('%Y-%m-%d %H:%M')}  •  Powered by Groq (Llama 3)"
            ws[f"A{last}"].font      = Font(name="Arial", color="AAAAAA", size=8, italic=True)
            ws[f"A{last}"].alignment = center
            ws.row_dimensions[last].height = 14

            # freeze header rows
            ws.freeze_panes = "A7"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        # collect ai_text if it was generated
        ai_text_for_xlsx = None
        if use_ai and "AI Analysis:" in report_text:
            ai_text_for_xlsx = report_text.split("**AI Analysis:**")[-1].strip()

        xlsx_buf = build_report_xlsx(
            row, recovery, head_grade, feed_rate, sipx, frother, ph,
            selected_shift_label, view, status, ai_text=ai_text_for_xlsx
        )
        st.download_button(
            "⬇️ Download Report (.xlsx)",
            xlsx_buf,
            file_name=f"shift_report_{selected_shift_label.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: ASK THE PLANT
# ─────────────────────────────────────────────────────────────────────────────
elif page == "💬 Ask the Plant":
    st.markdown('<div class="section-header">💬 Ask the Plant</div>', unsafe_allow_html=True)
    st.markdown("Natural language Q&A over your shift history. Powered by Groq (Llama 3).")

    if df_data is None:
        st.error("Dataset not loaded.")
        st.stop()

    df = df_data.copy()

    # Build context summary from dataset
    @st.cache_data
    def build_context(df):
        n = len(df)
        avg_rec = df[TARGET].mean() if TARGET in df.columns else "N/A"
        date_range = f"{df['Date'].min().date()} to {df['Date'].max().date()}" if "Date" in df.columns else "unknown"
        cols = ", ".join(df.columns.tolist()[:30])
        ctx = (
            f"You are an AI assistant helping plant operators analyze copper flotation data. "
            f"The dataset has {n} shift records from {date_range}. "
            f"Average recovery is {avg_rec:.2f}%. Available columns include: {cols}. "
            f"Features used in the ML model: {', '.join(FEATURES[:10])} and {len(FEATURES)-10} more. "
            f"Target is Recovery (%), with a plant target of 85%. "
            f"Answer operator questions clearly. For specific numbers, try to use the data. "
            f"If asked for calculations, provide them numerically."
        )
        return ctx

    system_ctx = build_context(df)

    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    for msg in st.session_state.chat_history:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    # Suggested questions
    if not st.session_state.chat_history:
        st.markdown("**Suggested questions:**")
        suggestions = [
            "What was the average recovery last month?",
            "Which shift has the highest average recovery?",
            "What features most affect recovery?",
            "How many shifts had recovery below 80%?",
        ]
        cols = st.columns(2)
        for i, s in enumerate(suggestions):
            if cols[i % 2].button(s, key=f"sug_{i}"):
                st.session_state.chat_history.append({"role": "user", "content": s})
                st.rerun()

    if prompt := st.chat_input("Ask about your plant data..."):
        st.session_state.chat_history.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        # Try to answer from data directly for simple queries
        data_answer = None
        pl = prompt.lower()
        if "average recovery" in pl or "avg recovery" in pl:
            data_answer = f"The average recovery across all {len(df)} shifts is **{df[TARGET].mean():.2f}%**."
        elif "below 80" in pl or "under 80" in pl:
            n_below = (df[TARGET] < 80).sum()
            data_answer = f"**{n_below}** shifts ({n_below/len(df)*100:.1f}%) had recovery below 80%."
        elif "above 85" in pl or "target" in pl:
            n_above = (df[TARGET] >= 85).sum()
            data_answer = f"**{n_above}** shifts ({n_above/len(df)*100:.1f}%) met or exceeded the 85% target."

        if data_answer:
            response = data_answer
        else:
            try:
                from groq import Groq
                groq_client = Groq(api_key=os.environ.get("GROQ_API_KEY", ""))
                messages = [{"role": "system", "content": system_ctx}] + \
                           [{"role": m["role"], "content": m["content"]} for m in st.session_state.chat_history]
                chat_completion = groq_client.chat.completions.create(
                    messages=messages,
                    model="llama-3.3-70b-versatile",
                    max_tokens=500,
                )
                response = chat_completion.choices[0].message.content
                if not response:
                    response = "Sorry, I couldn't get a response. Please check your GROQ_API_KEY in the .env file."
            except Exception as e:
                response = f"Groq API call failed: {e}. Make sure GROQ_API_KEY is set in your .env file."

        st.session_state.chat_history.append({"role": "assistant", "content": response})
        with st.chat_message("assistant"):
            st.markdown(response)

    if st.button("🗑️ Clear Chat"):
        st.session_state.chat_history = []
        st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: RECOVERY TREND FORECASTER
# ─────────────────────────────────────────────────────────────────────────────
elif page == "📈 Recovery Trend Forecaster":
    st.markdown('<div class="section-header">📈 Recovery Trend Forecaster</div>', unsafe_allow_html=True)
    st.markdown("7-day rolling recovery forecast based on planned shift inputs.")

    if not models:
        st.error("No models loaded.")
        st.stop()

    model_choice = st.selectbox("Forecast Model", list(models.keys()))
    model = models[model_choice]

    st.markdown("#### Enter 7-Day Planned Shift Parameters")
    st.markdown("Fill in the expected values for each of the next 7 days (3 shifts/day).")

    fc_bm1, fc_bm2 = st.columns(2)
    forecast_ball_mills = fc_bm1.number_input(
        "Ball Mills Running (applies to all 7 days)",
        value=7, step=1, min_value=1, max_value=20,
        help="Each ball mill consumes ~11 lime bags on average"
    )
    forecast_lime_bags = forecast_ball_mills * 11
    fc_bm2.info(f"🪨 Lime Bags per shift (auto): **{forecast_lime_bags}**  ({forecast_ball_mills} mills × 11)")

    forecast_rows = []
    for day_offset in range(7):
        date = datetime.today() + timedelta(days=day_offset)
        st.markdown(f"**Day {day_offset+1} — {date.strftime('%a %d %b')}**")
        dc1, dc2, dc3, dc4 = st.columns(4)
        hg    = dc1.number_input(f"Head Grade D{day_offset+1}",    0.1,  5.0, 1.5,   step=0.05, key=f"hg_{day_offset}")
        fr    = dc2.number_input(f"Feed Rate D{day_offset+1}",    20.0,200.0, 85.0,  step=1.0,  key=f"fr_{day_offset}")
        sipx  = dc3.number_input(f"SIPX D{day_offset+1}",          0.0, 50.0, 18.0,  step=0.5,  key=f"sipx_{day_offset}")
        ph    = dc4.number_input(f"pH D{day_offset+1}",            7.0, 13.0, 10.5,  step=0.1,  key=f"ph_{day_offset}")
        for shift in [1, 2, 3]:
            row = {
                "date": date.date(),
                "shift": shift,
                "Ore Milled (MT)": 800,
                "Head Grade (%Cu)": hg,
                "COPPER IN HEAD (MT)": hg * 800 * 0.01,
                "Feed Rate (MT/h)": fr,
                "Grinding kWh": 12000,
                "Lime Bags": forecast_lime_bags,
                "T Reagent (cc)": 1200,
                "Pine Oil (cc)": 400,
                "Flotation pH": ph,
                "Milling Running Hours": 8,
                "SIPX Dose (g/t)": sipx,
                "Frother Dose (g/t)": 10,
                "Depressant Dose (g/t)": 5,
                "Prev_Recovery (%)": 84,
                "Prev_Feed Rate (MT/h)": fr,
                "Prev_Head Grade (%Cu)": hg,
                "Prev_Flotation pH": ph,
                "Roll7_Recovery (%)": 83.5,
                "Roll7_Head Grade (%Cu)": hg,
                "Roll7_Feed Rate (MT/h)": fr,
                "Feed_Condition_Num": 0.0,
                "Shift_Num": float(shift),
                "Month": float(date.month),
                "Day_of_Week": float(date.weekday()),
            }
            forecast_rows.append(row)

    if st.button("📊 Generate 7-Day Forecast", use_container_width=True):
        forecast_df = pd.DataFrame(forecast_rows)
        Xf = forecast_df[FEATURES]
        forecast_df["Predicted Recovery (%)"] = model.predict(Xf)

        # Daily average
        daily = forecast_df.groupby("date")["Predicted Recovery (%)"].agg(["mean", "min", "max"]).reset_index()
        daily.columns = ["Date", "Avg Recovery", "Min Recovery", "Max Recovery"]

        st.markdown("#### Forecast Results")
        c1, c2, c3 = st.columns(3)
        c1.metric("7-Day Avg Forecast",   f"{daily['Avg Recovery'].mean():.2f}%")
        c2.metric("Best Day",             f"{daily['Avg Recovery'].max():.2f}%")
        c3.metric("Worst Day",            f"{daily['Avg Recovery'].min():.2f}%")

        fig, ax = plt.subplots(figsize=(12, 4))
        ax.plot(daily["Date"], daily["Avg Recovery"], "o-", color="#3b82f6", linewidth=2, label="Avg Predicted")
        ax.fill_between(daily["Date"], daily["Min Recovery"], daily["Max Recovery"], alpha=0.2, color="#3b82f6", label="Min–Max Range")
        ax.axhline(85, color="#e63946", linestyle="--", linewidth=1.5, label="Target 85%")
        ax.set_xlabel("Date")
        ax.set_ylabel("Predicted Recovery (%)")
        ax.set_title("7-Day Recovery Forecast")
        ax.legend()
        ax.grid(alpha=0.2)
        plt.tight_layout()
        st.pyplot(fig)
        plt.close()

        st.markdown("#### Daily Summary Table")
        st.dataframe(daily.style.format({"Avg Recovery": "{:.2f}%", "Min Recovery": "{:.2f}%", "Max Recovery": "{:.2f}%"}), use_container_width=True)

        # Export
        csv = daily.to_csv(index=False)
        st.download_button("⬇️ Download Forecast CSV", csv, file_name="recovery_forecast_7day.csv", mime="text/csv")
